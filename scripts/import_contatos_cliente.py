#!/usr/bin/env python3
"""
import_contatos_cliente.py — importa aba Contatos da planilha
Contatos-Clientes-Larissa.xlsx pra tabela public.contatos_cliente.

Estrutura esperada (aba "Contatos"):
  tipo | identificador | documento_cliente | nome_pessoa | cargo | observacao | ativo
  + 2 colunas novas que Larissa precisa adicionar:
  ordem | tipo_uso

Uso:
  set -a && source .env.local && set +a
  python3 scripts/import_contatos_cliente.py
"""

import json
import os
import sys
from pathlib import Path

import openpyxl
import requests


SUPABASE_URL = os.environ["SUPABASE_URL"].rstrip("/")
SERVICE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
PLANILHA = Path.home() / "Downloads" / "Contatos-Clientes-Larissa.xlsx"
ABA = "Contatos"


def parse_planilha():
    if not PLANILHA.exists():
        print(f"✕ Planilha não achada: {PLANILHA}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(PLANILHA, data_only=True)
    if ABA not in wb.sheetnames:
        print(f"✕ Aba '{ABA}' não encontrada. Abas: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[ABA]
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().lower() if c else "" for c in row]
            continue
        if i == 1 and (row[0] is None or "ex." in str(row[0]).lower() or "WhatsApp:" in str(row[0]) or "(11)" in str(row[0])):
            continue  # linha de exemplo/instrução
        if not any(row):
            continue

        record = dict(zip(headers, row))

        tipo = (record.get("tipo") or "").strip().lower()
        if tipo not in ("email", "whatsapp", "dominio"):
            continue

        identificador = str(record.get("identificador") or "").strip()
        # Excel guarda CNPJ como float quando célula é numérica → str() vira
        # "71336101000186.0", e o filtro de dígitos pega o ".0" como "0"
        # gerando CNPJ de 15 dígitos. Trata int/float ANTES de virar string.
        raw_doc = record.get("documento_cliente")
        if isinstance(raw_doc, (int, float)) and raw_doc:
            documento = str(int(raw_doc))
        else:
            documento = str(raw_doc or "").strip()
        documento = "".join(c for c in documento if c.isdigit())

        if not identificador or not documento:
            continue

        ativo_raw = str(record.get("ativo") or "sim").strip().lower()
        ativo = ativo_raw in ("sim", "yes", "true", "1")

        ordem = record.get("ordem")
        try:
            ordem = int(ordem) if ordem else 1
        except (TypeError, ValueError):
            ordem = 1

        tipo_uso = (record.get("tipo_uso") or "geral").strip().lower()
        if tipo_uso not in ("geral", "cobranca", "logistico", "financeiro", "comercial"):
            tipo_uso = "geral"

        rows.append({
            "documento_cliente": documento,
            "tipo": tipo,
            "identificador": identificador,
            "ordem": ordem,
            "tipo_uso": tipo_uso,
            "nome_pessoa": (record.get("nome_pessoa") or None),
            "cargo": (record.get("cargo") or None),
            "observacao": (record.get("observacao") or None),
            "ativo": ativo,
        })

    return rows


def upsert(rows):
    url = f"{SUPABASE_URL}/rest/v1/contatos_cliente"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
    }
    BATCH = 100
    total = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        res = requests.post(url, headers=headers, data=json.dumps(batch), timeout=30)
        if not res.ok:
            print(f"  ✕ batch {i}: {res.status_code} {res.text[:300]}", file=sys.stderr)
            sys.exit(1)
        total += len(batch)
    return total


def main():
    print(f"→ Lendo: {PLANILHA}")
    rows = parse_planilha()
    if not rows:
        print("⚠ Nenhum contato válido encontrado.", file=sys.stderr)
        sys.exit(1)

    por_tipo = {}
    for r in rows:
        por_tipo[r["tipo"]] = por_tipo.get(r["tipo"], 0) + 1
    print(f"  Total: {len(rows)} contatos ({', '.join(f'{k}:{v}' for k, v in por_tipo.items())})")

    print(f"→ Inserindo em public.contatos_cliente...")
    n = upsert(rows)
    print(f"  ✓ {n} rows inseridas")

    print("\n✦ Pronto.")


if __name__ == "__main__":
    main()
