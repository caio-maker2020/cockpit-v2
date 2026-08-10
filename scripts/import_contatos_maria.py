#!/usr/bin/env python3
# =============================================================================
# import_contatos_maria.py — gera migration/2026-08-10_322_seed_contatos_maria.sql
# a partir da planilha "Contatos clientes - Operador e transportador (1) (1).xlsx"
# (padrão do import_relacionamento_atualizado.py: staging + guardas + camadas).
#
# Entendimento TRAVADO com o Caio em 10/08 (plano aprovado):
#  * Aba 1 seção AGV: contato é por REMETENTE e vale pros 5 CNPJs pagadores AGV
#    ("o mesmo remetente pode sair como vários CNPJs pagadores da AGV").
#  * Aba 1 seção 3PL (LEXMARK/CTIS/ITAÚ): vale só pro pagador 3PL.
#  * Correções: OURO FINO = só 20258278000685; GLENMARK = 44363661000580;
#    VIRBAC/SYNTEC = dois remetentes (56921166000790 + 02177011000177).
#  * ordem = posição no bloco (1ª pessoa = TO default).
#  * Aba 2: contatos GERAIS (cnpj_remetente NULL); linha sem pagador = contato
#    extra do bloco imediatamente acima (caso Penske/Lucas).
# =============================================================================
import openpyxl
import re
import sys
from collections import OrderedDict

XLSX = "/Users/caiodevasconcelos/Downloads/Contatos clientes - Operador e transportador (1) (1).xlsx"
OUT = "migration/2026-08-10_322_seed_contatos_maria.sql"

AGV_PAGADORES = [
    ("02905424001879", "AGV LOG SA VINHEDO (A3)"),
    ("02905424010355", "AGV LOGISTICA S.A (A.)"),
    ("02905424001283", "AGV LOGISTICA S/A (A.)"),
    ("02905424006919", "AGV LOGISTICA S.A (A.)"),
    ("02905424002760", "AGV LOGISTICA SA (..)"),
]
PAG_3PL = ("23429671000178", "3PL BRASIL LOGISTICA S.A. (A.)")
REMETENTES_3PL = {"00767378000115", "01644731003743", "60701190000104"}
GLENMARK_CNPJ = "44363661000580"
OURO_FINO_CNPJ = "20258278000685"


def dig(v):
    return re.sub(r"\D", "", str(v)) if v is not None else ""


def esc(s):
    return str(s).replace("'", "''").strip()


def remetentes_da_celula(raw_txt, nome_bloco):
    nome_up = (nome_bloco or "").upper()
    if "OURO FINO" in nome_up:
        return [OURO_FINO_CNPJ]
    if "GLENMARK" in nome_up:
        return [GLENMARK_CNPJ]
    txt = str(raw_txt or "")
    if "/" in txt:
        return [dig(x).zfill(14) for x in txt.split("/") if dig(x)]
    d = dig(txt)
    return [d.zfill(14)] if d else []


def parse_aba1(wb):
    ws = wb["Contatos AGV"]
    # expande merges das colunas A (cnpj rem) e B (nome rem) — a autoridade do bloco
    remC, remN = {}, {}
    for rng in ws.merged_cells.ranges:
        top = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            if rng.min_col == 1 and rng.max_col == 1:
                remC[r] = top
            if rng.min_col == 2 and rng.max_col == 2:
                remN[r] = top
    for row in ws.iter_rows(min_row=2, max_col=2):
        for c in row:
            if c.value not in (None, "") and c.row not in (remC if c.column == 1 else remN):
                (remC if c.column == 1 else remN)[c.row] = c.value
    # carry-down: a mescla de alguns blocos NÃO cobre todas as linhas de
    # contato (folga de layout — ex: bloco AVERT). O remetente vale até o
    # próximo bloco começar (estrutura visual confirmada com o Caio).
    ult_c, ult_n = None, None
    for r in range(2, ws.max_row + 1):
        if r in remC and str(remC[r]).strip():
            ult_c = remC[r]
        elif ult_c is not None:
            remC[r] = ult_c
        if r in remN and str(remN[r]).strip():
            ult_n = remN[r]
        elif ult_n is not None:
            remN[r] = ult_n

    # blocos na ordem da planilha: (remetentes[], nome, [(pessoa, email, tel), ...])
    blocos = OrderedDict()
    for row in ws.iter_rows(min_row=2, max_col=5):
        r = row[0].row
        email = str(row[4].value or "").strip().lower()
        if "@" not in email:
            continue
        nome_bloco = str(remN.get(r) or "").strip()
        rems = remetentes_da_celula(remC.get(r), nome_bloco)
        if not rems:
            print(f"ABORT: linha {r} tem e-mail {email} sem remetente resolvido", file=sys.stderr)
            sys.exit(1)
        chave = (tuple(rems), nome_bloco)
        blocos.setdefault(chave, [])
        pessoa = esc(row[2].value or "")
        tel = esc(row[3].value or "")
        blocos[chave].append((pessoa, email, tel))
    return blocos


def parse_aba2(wb):
    ws = wb["Demais clientes"]
    out, atual = [], None
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        cnpj, cliente, resp, tel, email = [("" if v is None else str(v).strip()) for v in row]
        email = email.lower()
        if dig(cnpj):
            atual = (dig(cnpj).zfill(14), cliente)
        if "@" not in email:
            continue
        if atual is None:
            print(f"ABORT: aba 2 tem contato {email} antes de qualquer pagador", file=sys.stderr)
            sys.exit(1)
        out.append((atual[0], atual[1], esc(resp), email, esc(tel)))
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    blocos = parse_aba1(wb)
    aba2 = parse_aba2(wb)

    # materializa: (pagador, pagador_nome, remetente_cnpj, remetente_nome, ordem, pessoa, email, tel)
    linhas = []
    for (rems, nome_bloco), contatos in blocos.items():
        eh_3pl = all(r in REMETENTES_3PL for r in rems)
        pagadores = [PAG_3PL] if eh_3pl else AGV_PAGADORES
        for rem in rems:
            for pag_c, pag_n in pagadores:
                for ordem, (pessoa, email, tel) in enumerate(contatos, 1):
                    linhas.append((pag_c, pag_n, rem, nome_bloco, ordem, pessoa, email, tel))
    for pag_c, pag_n, pessoa, email, tel in [(a, b, c, d, e) for a, b, c, d, e in aba2]:
        pass  # (só pra clareza; aba2 tratada abaixo)

    aba2_linhas = []
    ordem_por_pag = {}
    for pag_c, pag_n, pessoa, email, tel in aba2:
        ordem_por_pag[pag_c] = ordem_por_pag.get(pag_c, 0) + 1
        aba2_linhas.append((pag_c, pag_n, ordem_por_pag[pag_c], pessoa, email, tel))

    n_agv = sum(1 for l in linhas if l[0] != PAG_3PL[0])
    n_3pl = sum(1 for l in linhas if l[0] == PAG_3PL[0])
    pag_especificos = sorted({l[0] for l in linhas})
    print(f"blocos de remetente: {len(blocos)} | linhas específicas: {len(linhas)} (AGV {n_agv} + 3PL {n_3pl}) | aba2 gerais: {len(aba2_linhas)}")

    values_esp = ",\n".join(
        f"('{p}','{esc(pn)}','{r}','{esc(rn)}',{o},'{esc(pe)}','{esc(em)}','{esc(te)}')"
        for p, pn, r, rn, o, pe, em, te in linhas
    )
    values_ger = ",\n".join(
        f"('{p}','{esc(pn)}',{o},'{esc(pe)}','{esc(em)}','{esc(te)}')"
        for p, pn, o, pe, em, te in aba2_linhas
    )
    pag_list = ", ".join(f"'{p}'" for p in sorted({l[0] for l in linhas} | {a[0] for a in aba2_linhas}))

    sql = f"""-- ============================================================================
-- 2026-08-10_322 — SEED dos contatos da MARIA (planilha 10/08; GERADA por
-- scripts/import_contatos_maria.py — NÃO editar à mão, regenerar).
-- Aba 1: contatos POR REMETENTE (AGV ×5 filiais + 3PL). Aba 2: gerais.
-- Correções do Caio aplicadas: OURO FINO/GLENMARK/VIRBAC-SYNTEC (ver script).
-- ============================================================================
BEGIN;

CREATE TEMP TABLE _esp (
  pagador text, pagador_nome text, remetente text, remetente_nome text,
  ordem int, pessoa text, email text, tel text
) ON COMMIT DROP;
INSERT INTO _esp VALUES
{values_esp};

CREATE TEMP TABLE _ger (
  pagador text, pagador_nome text, ordem int, pessoa text, email text, tel text
) ON COMMIT DROP;
INSERT INTO _ger VALUES
{values_ger};

-- ---- GUARDAS -----------------------------------------------------------
DO $g$
DECLARE v_esp int; v_ger int; v_maria uuid; v_dono text;
BEGIN
  SELECT count(*) INTO v_esp FROM _esp;
  SELECT count(*) INTO v_ger FROM _ger;
  IF v_esp <> {len(linhas)} OR v_ger <> {len(aba2_linhas)} THEN
    RAISE EXCEPTION 'G1: staging divergente (esp=% ger=%)', v_esp, v_ger;
  END IF;
  SELECT id INTO v_maria FROM operadores WHERE nome='MARIA';
  IF v_maria IS NULL THEN RAISE EXCEPTION 'G2: MARIA não existe'; END IF;
  -- 5º AGV não pode estar em carteira de OUTRO operador ativo (INV-036a)
  SELECT string_agg(nome, ', ') INTO v_dono FROM operadores o
   WHERE o.nome <> 'MARIA' AND o.ativo
     AND '02905424002760' = ANY(SELECT lpad(regexp_replace(x,'\\D','','g'),14,'0') FROM unnest(o.carteira) x);
  IF v_dono IS NOT NULL THEN RAISE EXCEPTION 'G3: 5º AGV em carteira de %', v_dono; END IF;
END $g$;

-- ---- Camada 1: 5º CNPJ AGV entra no catálogo e na carteira da MARIA ----
INSERT INTO clientes (cnpj_cpf, nome, segmento_codigo, segmento_nome, ativo)
VALUES ('02905424002760', 'AGV LOGISTICA SA', '040', 'OPERADOR LOGISTICO', true)
ON CONFLICT (cnpj_cpf) DO UPDATE SET segmento_codigo='040', segmento_nome='OPERADOR LOGISTICO', ativo=true;

UPDATE operadores
SET carteira = (SELECT array_agg(DISTINCT c) FROM unnest(carteira || '{{02905424002760}}'::text[]) c)
WHERE nome='MARIA'
  AND NOT ('02905424002760' = ANY(SELECT lpad(regexp_replace(x,'\\D','','g'),14,'0') FROM unnest(carteira) x));

-- ---- Camada 2: contatos (DELETE escopado + INSERT) ---------------------
DELETE FROM contatos_cliente
WHERE documento_cliente IN ({pag_list}) AND tipo='email';

INSERT INTO contatos_cliente
  (documento_cliente, tipo, identificador, ordem, tipo_uso, nome_pessoa, observacao, ativo, operador_responsavel_id, cnpj_remetente)
SELECT e.pagador, 'email', e.email, e.ordem, 'logistico', nullif(e.pessoa,''),
       'Remetente: '||e.remetente_nome||' ('||e.remetente||')'||CASE WHEN e.tel<>'' AND e.tel<>'-' THEN ' · tel '||e.tel ELSE '' END||' — planilha Maria 10/08',
       true, (SELECT id FROM operadores WHERE nome='MARIA'), e.remetente
FROM _esp e;

INSERT INTO contatos_cliente
  (documento_cliente, tipo, identificador, ordem, tipo_uso, nome_pessoa, observacao, ativo, operador_responsavel_id, cnpj_remetente)
SELECT g.pagador, 'email', g.email, g.ordem, 'logistico', nullif(g.pessoa,''),
       'Contato geral'||CASE WHEN g.tel<>'' AND g.tel<>'-' THEN ' · tel '||g.tel ELSE '' END||' — planilha Maria 10/08',
       true, (SELECT id FROM operadores WHERE nome='MARIA'), NULL
FROM _ger g;

-- domínio (whitelist inbound) pros pagadores novos que não têm
INSERT INTO contatos_cliente (documento_cliente, tipo, identificador, ordem, tipo_uso, ativo, operador_responsavel_id)
SELECT DISTINCT e.pagador, 'dominio', split_part(e.email,'@',2), 1, 'geral', true,
       (SELECT id FROM operadores WHERE nome='MARIA')
FROM (SELECT pagador, email FROM _esp UNION SELECT pagador, email FROM _ger) e
WHERE NOT EXISTS (
  SELECT 1 FROM contatos_cliente c
  WHERE c.documento_cliente=e.pagador AND c.tipo='dominio' AND c.identificador=split_part(e.email,'@',2)
);

-- ---- ASSERTS pós-seed --------------------------------------------------
DO $a$
DECLARE v text;
BEGIN
  -- ZOETIS resolve o contato certo em TODAS as 5 filiais AGV
  FOREACH v IN ARRAY ARRAY['02905424001879','02905424010355','02905424001283','02905424006919','02905424002760'] LOOP
    IF public.resolver_email_cobranca_cliente(v,'logistico','01770356000177') IS DISTINCT FROM 'luis.folli@agv.com.br' THEN
      RAISE EXCEPTION 'A1: ZOETIS na filial % não resolveu luis.folli', v;
    END IF;
  END LOOP;
  -- AGV SEM remetente → NULL (sem contato geral: modal exige escolha)
  IF public.resolver_email_cobranca_cliente('02905424001879','logistico') IS NOT NULL THEN
    RAISE EXCEPTION 'A2: AGV sem remetente deveria resolver NULL';
  END IF;
  -- 3PL: LEXMARK resolve cselexmark
  IF public.resolver_email_cobranca_cliente('23429671000178','logistico','00767378000115') IS DISTINCT FROM 'cselexmark@agv.com.br' THEN
    RAISE EXCEPTION 'A3: LEXMARK na 3PL não resolveu cselexmark';
  END IF;
  -- aba 2: geral continua funcionando (Penske)
  IF public.resolver_email_cobranca_cliente('65849838004278','logistico') IS DISTINCT FROM 'nelissa.barboza@penske.com.br' THEN
    RAISE EXCEPTION 'A4: Penske geral não resolveu nelissa';
  END IF;
END $a$;

COMMIT;
"""
    open(OUT, "w", encoding="utf-8").write(sql)
    print(f"gerada: {OUT} ({len(sql.splitlines())} linhas)")


if __name__ == "__main__":
    main()
