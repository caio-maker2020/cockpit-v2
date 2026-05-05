#!/usr/bin/env python3
"""
gen_planilha_contatos.py — Gera planilha modelo (.xlsx) pra Larissa cadastrar
clientes + contatos + senhas de rastreio em paralelo.

Saída: docs/exports/Contatos-Clientes-Larissa.xlsx
Abas: INSTRUÇÕES, Clientes, Contatos

Mapeamento pra banco (futuro):
  - Clientes  → public.tracking_credentials (documento, nome_amigavel, senha, ...)
  - Contatos  → public.contatos_clientes (tipo, identificador, cnpj_pagador, ...)
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = Path(__file__).resolve().parent.parent / "docs" / "exports" / "Contatos-Clientes-Larissa.xlsx"


# Cores Sal Express
SAL_RED = "C8102E"
SAL_RED_LIGHT = "F8D7DA"
HEADER_BG = "2C2C2C"
HEADER_FG = "FFFFFF"
EXAMPLE_BG = "FFF3CD"
GREY_BORDER = "B8B8B8"

thin_border = Border(
    left=Side(style="thin", color=GREY_BORDER),
    right=Side(style="thin", color=GREY_BORDER),
    top=Side(style="thin", color=GREY_BORDER),
    bottom=Side(style="thin", color=GREY_BORDER),
)


def style_header(cell):
    cell.font = Font(bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border


def style_example(cell):
    cell.fill = PatternFill("solid", fgColor=EXAMPLE_BG)
    cell.font = Font(italic=True, color="6C6C6C")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border


def style_title(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill("solid", fgColor=SAL_RED)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_instrucoes(wb: Workbook):
    ws = wb.create_sheet("INSTRUÇÕES", 0)

    ws.column_dimensions["A"].width = 110

    ws["A1"] = "Cockpit Sal Express — Planilha de Cadastro de Clientes e Contatos"
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 28

    blocos = [
        ("", ""),
        ("PARA QUE SERVE ESSA PLANILHA",
         "O Cockpit é o sistema de IA que vai responder automaticamente clientes da Sal Express "
         "no WhatsApp e email. Pra ele identificar QUEM mandou a mensagem, precisamos de duas coisas:\n"
         "  1) Cadastro do CLIENTE (CNPJ + senha do rastreio SSW)\n"
         "  2) Cadastro dos CONTATOS de cada cliente (telefones, emails, domínios)\n"
         "Quando chegar uma mensagem, o sistema casa o telefone/email com um contato cadastrado, "
         "descobre o CNPJ do cliente, e usa a senha do rastreio pra puxar status no SSW."),

        ("ORDEM DE PREENCHIMENTO",
         "1º — Preencha a aba 'Clientes' com o CNPJ e a senha de rastreio do SSW de cada cliente.\n"
         "2º — Preencha a aba 'Contatos' usando o mesmo CNPJ que você cadastrou em 'Clientes'.\n"
         "Cada cliente pode ter VÁRIOS contatos (uma linha por contato)."),

        ("REGRAS DE FORMATO",
         "• CNPJ/CPF: SEM pontos, traços ou barras. Só números. Ex.: 12345678000199\n"
         "• Telefone WhatsApp: SEMPRE com DDI 55 + DDD + número, só dígitos. Ex.: 5531999999999\n"
         "• Email: completo, minúsculas. Ex.: compras@cliente.com.br\n"
         "• Domínio: SEM o @, só o domínio. Ex.: cliente.com.br (vale pra qualquer email @cliente.com.br)\n"
         "• Senha de rastreio SSW: exatamente como aparece no painel SSW (case sensitive)\n"
         "• Ativo: 'sim' ou 'não' (default: sim)"),

        ("TIPOS DE CONTATO (coluna 'tipo' na aba Contatos)",
         "• whatsapp     — número de telefone individual (uma pessoa específica)\n"
         "• email        — email individual completo (uma pessoa específica)\n"
         "• dominio      — domínio inteiro (qualquer pessoa @aquele.dominio cai no mesmo cliente)\n"
         "Use 'dominio' quando o cliente tem muitas pessoas que mandam emails — economiza cadastro."),

        ("SENHA DO RASTREIO SSW — IMPORTANTE",
         "Cada cliente tem uma senha pra acessar o rastreio dele no SSW. É essa senha que o Cockpit "
         "usa pra puxar o status das cargas via API.\n"
         "• Se o cliente NÃO tem senha cadastrada no SSW, peça pro time operacional gerar uma.\n"
         "• Se o cliente já é atendido há tempos, provavelmente a senha já existe — confirma com o "
         "comercial ou no próprio painel SSW (módulo de cadastro de pagador).\n"
         "• Cliente sem senha: deixa a coluna em branco e marca observação 'sem senha — pendente'."),

        ("DICAS PRA NÃO ERRAR",
         "• Não duplique linhas. Um CNPJ por linha em Clientes; um contato (tel/email/domínio) por linha em Contatos.\n"
         "• Se o mesmo telefone serve pra 2 clientes diferentes (raro), cadastre 2 vezes — mas avisa o Caio antes.\n"
         "• Se não souber algo (nome da pessoa, cargo), deixa em branco. NÃO inventa.\n"
         "• Use a coluna 'observacao' pra qualquer detalhe importante (ex.: 'só atende manhã', 'gerente novo')."),

        ("EXEMPLOS PREENCHIDOS",
         "Nas abas Clientes e Contatos, as primeiras 3 linhas são EXEMPLOS (fundo amarelo). "
         "Use elas como referência e DEPOIS APAGUE antes de mandar de volta."),

        ("DÚVIDAS", "Fala com o Caio (caio@salexpress.com.br) — qualquer dúvida, manda no WhatsApp dele."),
    ]

    row = 3
    for titulo, texto in blocos:
        if titulo:
            ws.cell(row=row, column=1, value=titulo).font = Font(bold=True, color=SAL_RED, size=12)
            ws.row_dimensions[row].height = 18
            row += 1
        if texto:
            ws.cell(row=row, column=1, value=texto).alignment = Alignment(wrap_text=True, vertical="top")
            # altura proporcional ao número de linhas
            n_linhas = texto.count("\n") + 1
            ws.row_dimensions[row].height = max(20, n_linhas * 16)
            row += 2


def write_clientes(wb: Workbook):
    ws = wb.create_sheet("Clientes")

    headers = [
        ("documento", 18, "CPF (11) ou CNPJ (14) do cliente. SÓ NÚMEROS."),
        ("nome_cliente", 35, "Nome curto do cliente (como você chama ele no dia a dia)."),
        ("senha_tracking_ssw", 20, "Senha de rastreio do SSW. Case sensitive."),
        ("segmento", 18, "Ex.: industria, varejo, atacado, e-commerce. (opcional)"),
        ("responsavel_relacionamento", 25, "Quem da Sal Express atende esse cliente. (opcional)"),
        ("observacao", 35, "Qualquer informação extra. (opcional)"),
        ("ativo", 8, "sim ou não. Default: sim."),
    ]

    for col_idx, (name, width, _help) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Tooltips na linha 2 (em italic, fundo cinza claro) — descrevem o campo
    for col_idx, (_name, _width, helptxt) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=helptxt)
        cell.font = Font(italic=True, color="6C6C6C", size=9)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border
    ws.row_dimensions[2].height = 40

    # Exemplos
    exemplos = [
        ("12345678000199", "Indústria ABC Ltda", "abc2025!", "industria", "João (comercial)", "Cliente desde 2018", "sim"),
        ("98765432000155", "Varejo XPTO", "xpto-rastreio", "varejo", "Maria", "Atende WhatsApp e email", "sim"),
        ("11122233344", "João da Silva ME", "joao123", "atacado", "", "Cliente novo — abril/2026", "sim"),
    ]

    for i, ex in enumerate(exemplos, start=3):
        for col_idx, val in enumerate(ex, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            style_example(cell)

    # Validação: ativo (sim/não)
    dv_ativo = DataValidation(type="list", formula1='"sim,não"', allow_blank=True)
    dv_ativo.add(f"G3:G2000")
    ws.add_data_validation(dv_ativo)

    # Linhas em branco prontas pra preencher (até 200)
    for r in range(6, 200):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = thin_border

    ws.freeze_panes = "A3"


def write_contatos(wb: Workbook):
    ws = wb.create_sheet("Contatos")

    headers = [
        ("tipo", 12, "whatsapp, email, ou dominio."),
        ("identificador", 32, "WhatsApp: 5531999999999 | Email: pessoa@cliente.com.br | Domínio: cliente.com.br"),
        ("documento_cliente", 18, "CNPJ/CPF do cliente (igual ao da aba Clientes). SÓ NÚMEROS."),
        ("nome_pessoa", 25, "Nome de quem manda mensagem. (opcional pra dominio)"),
        ("cargo", 20, "Ex.: comprador, gerente. (opcional)"),
        ("observacao", 35, "Detalhes importantes. (opcional)"),
        ("ativo", 8, "sim ou não. Default: sim."),
    ]

    for col_idx, (name, width, _) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx, (_n, _w, helptxt) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=helptxt)
        cell.font = Font(italic=True, color="6C6C6C", size=9)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border
    ws.row_dimensions[2].height = 40

    exemplos = [
        ("whatsapp", "5531999998888", "12345678000199", "João Silva", "Comprador", "Atende manhã", "sim"),
        ("email", "joao@abc.com.br", "12345678000199", "João Silva", "Comprador", "", "sim"),
        ("dominio", "abc.com.br", "12345678000199", "", "", "Vale pra qualquer email @abc.com.br", "sim"),
        ("whatsapp", "5511988887777", "98765432000155", "Maria Santos", "Gerente compras", "", "sim"),
        ("dominio", "xpto.com.br", "98765432000155", "", "", "Pega todos os emails do XPTO", "sim"),
    ]

    for i, ex in enumerate(exemplos, start=3):
        for col_idx, val in enumerate(ex, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            style_example(cell)

    # Validações
    dv_tipo = DataValidation(type="list", formula1='"whatsapp,email,dominio"', allow_blank=False)
    dv_tipo.add("A3:A2000")
    ws.add_data_validation(dv_tipo)

    dv_ativo = DataValidation(type="list", formula1='"sim,não"', allow_blank=True)
    dv_ativo.add("G3:G2000")
    ws.add_data_validation(dv_ativo)

    # Bordas em linhas em branco
    for r in range(8, 500):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = thin_border

    ws.freeze_panes = "A3"


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_instrucoes(wb)
    write_clientes(wb)
    write_contatos(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK: {OUT}")


if __name__ == "__main__":
    main()
