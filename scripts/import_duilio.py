#!/usr/bin/env python3
"""
import_duilio.py — Importa clientes + contatos do operador Duilio Rafael.

Diferenças vs import_clientes_larissa.py (planilha por código SSW em abas):
- Planilha tem 1 aba só ("Clientes") com tudo
- Linha 0 = header, linha 1 = descrições/instruções (skip), linhas 2+ = dados
- Colunas: A=CNPJ, C=nome_cliente, D=segmento (texto livre — ignoramos),
           E=email, F=responsavel_relacionamento (interno), G=whatsapp
- Segmento SSW = "014" (Caio 2026-05-14, código fornecido).

Faz 3 coisas:
1. UPSERT em `clientes` (31 CNPJs únicos com nome + segmento_codigo='014')
2. UPSERT em `contatos_cliente` (44 linhas com email)
3. UPDATE `operadores` SET carteira/segmentos WHERE nome='DUILIO'

Uso:
  set -a && source .env.local && set +a
  python3 scripts/import_duilio.py
"""

import json
import os
import re
import sys
from pathlib import Path

import openpyxl
import requests


SUPABASE_URL = os.environ["SUPABASE_URL"].rstrip("/")
SERVICE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
PLANILHA = Path.home() / "Downloads" / "contatos-duilio Finalizado.xlsx"
ABA = "Clientes"

SEGMENTO_CODIGO = "014"
SEGMENTO_NOME = "FERRAMENTAS E CONSTRUCAO"
OPERADOR_NOME = "DUILIO"


def norm_cnpj(raw) -> str | None:
    if raw is None:
        return None
    # Excel guarda CNPJ como float às vezes — trata antes de str()
    if isinstance(raw, float):
        raw = int(raw)
    s = re.sub(r"\D", "", str(raw))
    # Excel também perde zeros à esquerda quando o CNPJ vira numero. Pad
    # pra 14 dígitos quando vier com 12 ou 13 (CNPJ válido sempre tem 14).
    # CPF (11 dígitos) também pode perder — pad pra 11 se vier com 9-10.
    if 12 <= len(s) <= 13:
        s = s.zfill(14)
    elif 9 <= len(s) <= 10:
        s = s.zfill(11)
    if len(s) in (11, 14):
        return s
    return None


def parse_planilha():
    """Retorna (clientes, contatos):
    clientes = list of {cnpj_cpf, nome, segmento_codigo, segmento_nome, ativo}
    contatos = list of {tipo, identificador, documento_cliente, nome_pessoa, ativo}
    """
    if not PLANILHA.exists():
        print(f"✕ Planilha não achada: {PLANILHA}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(PLANILHA, data_only=True)
    if ABA not in wb.sheetnames:
        print(f"✕ Aba '{ABA}' não encontrada. Abas: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[ABA]
    clientes_by_cnpj: dict[str, dict] = {}
    contatos: list[dict] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1:
            continue  # header + linha de instrução
        if not row or not row[0]:
            continue

        cnpj = norm_cnpj(row[0])
        if not cnpj:
            print(f"  ⚠ Linha {i+1}: cnpj inválido '{row[0]}' — pulado", file=sys.stderr)
            continue

        nome_pessoa = str(row[2]).strip() if len(row) > 2 and row[2] else None
        email = str(row[4]).strip() if len(row) > 4 and row[4] else None
        whats = str(row[6]).strip() if len(row) > 6 and row[6] else None
        # Caio 2026-05-15: col H ('Cliente') é o NOME DA EMPRESA (CORTAG, OVD, FG).
        # Col C é o nome da PESSOA contato. clientes.nome = empresa.
        nome_empresa = str(row[7]).strip() if len(row) > 7 and row[7] else None
        nome_cli_final = nome_empresa or nome_pessoa or "(sem nome)"

        # Cliente: dedup por CNPJ (mantém nome da empresa se houver)
        if cnpj not in clientes_by_cnpj:
            clientes_by_cnpj[cnpj] = {
                "cnpj_cpf": cnpj,
                "nome": nome_cli_final,
                "segmento_codigo": SEGMENTO_CODIGO,
                "segmento_nome": SEGMENTO_NOME,
                "ativo": True,
            }
        elif nome_empresa and not clientes_by_cnpj[cnpj]["nome"].strip() in (nome_empresa,):
            # Atualiza se primeira ocorrência veio sem nome de empresa
            clientes_by_cnpj[cnpj]["nome"] = nome_empresa

        # Contatos: 1 row por (cnpj, identificador, tipo). nome_pessoa = pessoa
        # individual (col C), não a empresa.
        nome_contato = nome_pessoa or nome_empresa or ""
        if email and "@" in email:
            contatos.append({
                "tipo": "email",
                "identificador": email.lower(),
                "documento_cliente": cnpj,
                "nome_pessoa": nome_contato,
                "ativo": True,
            })
        if whats:
            whats_digits = re.sub(r"\D", "", whats)
            if len(whats_digits) >= 10:
                contatos.append({
                    "tipo": "whatsapp",
                    "identificador": whats_digits,
                    "documento_cliente": cnpj,
                    "nome_pessoa": nome_contato,
                    "ativo": True,
                })

    return list(clientes_by_cnpj.values()), contatos


def buscar_operador_id(nome: str) -> str:
    url = f"{SUPABASE_URL}/rest/v1/operadores?nome=eq.{nome}&select=id"
    headers = {"apikey": SERVICE_KEY, "Authorization": f"Bearer {SERVICE_KEY}"}
    res = requests.get(url, headers=headers, timeout=30)
    if not res.ok or not res.json():
        print(f"  ✕ Operador {nome} não encontrado", file=sys.stderr)
        sys.exit(1)
    return res.json()[0]["id"]


def deletar_contatos_do_operador(operador_id: str) -> int:
    """Idempotência: apaga contatos antigos do operador antes do INSERT.
    Necessário porque contatos_cliente não tem unique constraint em (tipo, identificador)."""
    url = f"{SUPABASE_URL}/rest/v1/contatos_cliente?operador_responsavel_id=eq.{operador_id}"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Prefer": "return=representation",
    }
    res = requests.delete(url, headers=headers, timeout=30)
    if not res.ok:
        print(f"  ✕ DELETE contatos falhou: HTTP {res.status_code} {res.text[:300]}", file=sys.stderr)
        sys.exit(1)
    return len(res.json()) if res.text else 0


def upsert(table: str, rows: list[dict], on_conflict: str | None = None) -> int:
    if not rows:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if on_conflict:
        url += f"?on_conflict={on_conflict}"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    BATCH = 100
    inserted = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        res = requests.post(url, headers=headers, data=json.dumps(batch), timeout=30)
        if not res.ok:
            print(f"  ✕ Falhou batch {i} em {table}: HTTP {res.status_code} {res.text[:300]}", file=sys.stderr)
            sys.exit(1)
        inserted += len(batch)
    return inserted


def vincular_operador(clientes: list[dict]) -> tuple[list[str], list[str]]:
    cnpjs = sorted({c["cnpj_cpf"] for c in clientes})
    segmentos = sorted({c["segmento_codigo"] for c in clientes})

    url = f"{SUPABASE_URL}/rest/v1/operadores"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    res = requests.patch(
        f"{url}?nome=eq.{OPERADOR_NOME}",
        headers=headers,
        data=json.dumps({"carteira": cnpjs, "segmentos": segmentos}),
        timeout=30,
    )
    if not res.ok:
        print(f"  ✕ UPDATE {OPERADOR_NOME} falhou: HTTP {res.status_code} {res.text[:300]}", file=sys.stderr)
        sys.exit(1)
    return cnpjs, segmentos


def main():
    print(f"→ Lendo planilha: {PLANILHA}")
    clientes, contatos = parse_planilha()
    print(f"  ✓ {len(clientes)} clientes únicos")
    print(f"  ✓ {len(contatos)} contatos (email/whatsapp)")

    print(f"\n→ Buscando operador_id de {OPERADOR_NOME}...")
    operador_id = buscar_operador_id(OPERADOR_NOME)
    print(f"  ✓ {operador_id}")

    # Acrescenta operador_responsavel_id em cada contato (NOT NULL na tabela)
    for c in contatos:
        c["operador_responsavel_id"] = operador_id

    print(f"\n→ Upsert em public.clientes ({len(clientes)} rows)...")
    n = upsert("clientes", clientes)
    print(f"  ✓ {n} rows enviadas")

    # Caio 2026-05-15: aba CADASTROS Lovable lista rows de tracking_credentials.
    # Tracking público está deprecated (Fase 3) mas Lovable ainda usa essa
    # tabela como "lista de clientes do operador". Sem essa inserção, aba
    # aparece vazia mesmo com carteira populada. Senha NULL — SSW interno cobre.
    print(f"\n→ Upsert tracking_credentials stub pros {len(clientes)} clientes...")
    tcs = [{
        "documento": c["cnpj_cpf"],
        "nome_amigavel": c["nome"],
        "senha": None,
        "ativo": True,
        "operador_responsavel_id": operador_id,
        "notes": "Auto-gerado pelo import — tracking público deprecated; SSW interno cobre.",
    } for c in clientes]
    n = upsert("tracking_credentials", tcs)
    print(f"  ✓ {n} rows enviadas")

    # contatos_cliente: SEM unique em (tipo, identificador). Idempotência via
    # DELETE prévio dos contatos do operador → INSERT batch novo.
    print(f"\n→ DELETE contatos antigos do {OPERADOR_NOME} (idempotência)...")
    n_del = deletar_contatos_do_operador(operador_id)
    print(f"  ✓ {n_del} contatos antigos removidos")

    print(f"\n→ INSERT em public.contatos_cliente ({len(contatos)} rows)...")
    n = upsert("contatos_cliente", contatos)
    print(f"  ✓ {n} rows enviadas")

    print(f"\n→ Vinculando {OPERADOR_NOME}: carteira + segmentos...")
    cnpjs, segmentos = vincular_operador(clientes)
    print(f"  ✓ {OPERADOR_NOME}.carteira = {len(cnpjs)} CNPJs")
    print(f"  ✓ {OPERADOR_NOME}.segmentos = {segmentos}")

    print("\n✦ Pronto. NÃO ativa cockpit_ativo=true do Duilio — fazer manualmente em M6 (auth + Gmail).")


if __name__ == "__main__":
    main()
