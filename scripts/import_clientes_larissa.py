#!/usr/bin/env python3
"""
import_clientes_larissa.py — Importa clientes da planilha clientes_larissa.xlsx
pra tabela `clientes` do Cockpit. Cada aba é um segmento.

Uso: SUPABASE_URL=... SUPABASE_SERVICE_KEY=... python3 import_clientes_larissa.py
"""

import json
import os
import sys
from pathlib import Path
import openpyxl
import requests


SUPABASE_URL = os.environ["SUPABASE_URL"].rstrip("/")
SERVICE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
PLANILHA = Path.home() / "Downloads" / "clientes_larissa.xlsx"


def parse_planilha():
    """Lê todas as abas, retorna lista de dicts {cnpj_cpf, nome, segmento_codigo, segmento_nome}."""
    wb = openpyxl.load_workbook(PLANILHA, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        # "007 - MEDICAMENTO" → codigo='007', nome='MEDICAMENTO'
        parts = sheet_name.split(" - ", 1)
        codigo = parts[0].strip()
        nome_segmento = parts[1].strip() if len(parts) > 1 else sheet_name

        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if not row or not row[0]:
                continue
            cnpj = str(row[0]).strip()
            cnpj = "".join(c for c in cnpj if c.isdigit())  # só dígitos
            if len(cnpj) not in (11, 14):
                print(f"  ⚠ Pulou cnpj inválido na aba {sheet_name}: '{row[0]}' → '{cnpj}'", file=sys.stderr)
                continue
            nome_cliente = str(row[1]).strip() if len(row) > 1 and row[1] else "(sem nome)"
            out.append({
                "cnpj_cpf": cnpj,
                "nome": nome_cliente,
                "segmento_codigo": codigo,
                "segmento_nome": nome_segmento,
                "ativo": True,
            })
    return out


def upsert_clientes(rows):
    url = f"{SUPABASE_URL}/rest/v1/clientes"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates",
    }
    BATCH = 100
    inserted = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i+BATCH]
        res = requests.post(url, headers=headers, data=json.dumps(batch), timeout=30)
        if not res.ok:
            print(f"  ✕ Falhou batch {i}: HTTP {res.status_code} {res.text[:200]}", file=sys.stderr)
            sys.exit(1)
        inserted += len(batch)
    return inserted


def vincular_larissa(rows):
    """Atualiza operadores.carteira (CNPJs) + segmentos (códigos) da Larissa."""
    cnpjs = sorted({r["cnpj_cpf"] for r in rows})
    segmentos = sorted({r["segmento_codigo"] for r in rows})

    url = f"{SUPABASE_URL}/rest/v1/operadores"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    res = requests.patch(
        f"{url}?nome=eq.LARISSA",
        headers=headers,
        data=json.dumps({"carteira": cnpjs, "segmentos": segmentos}),
        timeout=30,
    )
    if not res.ok:
        print(f"  ✕ UPDATE Larissa falhou: HTTP {res.status_code} {res.text[:300]}", file=sys.stderr)
        sys.exit(1)
    return cnpjs, segmentos


def main():
    if not PLANILHA.exists():
        print(f"✕ Não achei planilha: {PLANILHA}", file=sys.stderr)
        sys.exit(1)

    print(f"→ Lendo planilha: {PLANILHA}")
    rows = parse_planilha()
    print(f"  Total: {len(rows)} clientes em {len({r['segmento_codigo'] for r in rows})} segmentos")

    print(f"→ Upserting {len(rows)} clientes em public.clientes...")
    n = upsert_clientes(rows)
    print(f"  ✓ {n} rows enviadas (merge em duplicatas)")

    print(f"→ Vinculando Larissa: {len(set(r['cnpj_cpf'] for r in rows))} CNPJs + segmentos...")
    cnpjs, segmentos = vincular_larissa(rows)
    print(f"  ✓ Larissa.carteira = {len(cnpjs)} CNPJs")
    print(f"  ✓ Larissa.segmentos = {segmentos}")

    print("\n✦ Pronto.")


if __name__ == "__main__":
    main()
